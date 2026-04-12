"""Ingest one-shot du fichier Bibles Greg.xlsx en index JSON structure.

Transforme la base metier Excel maintenue par le pathologiste en une
liste plate d'entrees typees (BiblesEntry) consommables par le pipeline
v4. Execution unique a chaque mise a jour du fichier source :

    python scripts/ingest_bibles_greg.py

Sortie : backend/retrieval/data/bibles_greg.json (engage dans le repo).

Convention : une fonction = une action. Zero logique metier dans la
fonction main, qui orchestre uniquement les etapes.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import TYPE_CHECKING

from openpyxl import load_workbook

REPO_ROOT: Path = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT / "backend"))

from schemas import BiblesEntry, Organe  # noqa: E402  # pyright: ignore[reportMissingImports]

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------------------------------------------------------
# Mapping feuille Excel -> organe canonique
# ---------------------------------------------------------------------------

SHEET_TO_ORGANE: dict[str, Organe] = {
    "DIG": "digestif",
    "BDig": "digestif",
    "URO": "urologie",
    "GYN": "gynecologie",
    "DERMATO": "dermatologie",
    "CR": "generic",
    "CYTO": "generic",
    "Cyto FCU": "gynecologie",
    "SEIN": "sein",
    "FOETO-P": "gynecologie",
    "ORL": "orl",
    "THO": "poumon",
    "ENDOC": "endocrinologie",
    "OS-ARTI": "os_articulations",
    "HEMATO": "hematologie",
    "T mous": "tissus_mous",
    "CV": "cardiovasculaire",
    "NEURO": "neurologie",
    "OPH": "ophtalmologie",
    "M": "generic",
}

# Feuilles a ignorer (templates orphelins, macros generiques)
SHEETS_TO_SKIP: set[str] = {"Feuil1"}

# Longueur minimale pour qu'un texte standard soit juge exploitable
MIN_TEXTE_LENGTH: int = 20


# ---------------------------------------------------------------------------
# Normalisation de texte
# ---------------------------------------------------------------------------


def normalize_text(raw: object) -> str:
    """Nettoie une valeur de cellule brute en str propre.

    Accepte ``object`` car openpyxl peut retourner str, int, float, Decimal,
    datetime, None ou des types formula exotiques. Conversion via ``str()``
    et nettoyage standard (strip, espaces multiples, lignes vides).
    """
    if raw is None:
        return ""
    text: str = str(raw).strip()
    lines: list[str] = [line.strip() for line in text.splitlines()]
    cleaned: list[str] = [line for line in lines if line]
    return "\n".join(cleaned)


# ---------------------------------------------------------------------------
# Detection du schema de colonnes par feuille
# ---------------------------------------------------------------------------


def detect_header_row(worksheet: "Worksheet") -> bool:
    """Detecte si la premiere ligne est un header (contient 'Topo' ou 'Lésion')."""
    row_iter = worksheet.iter_rows(min_row=1, max_row=1, values_only=True)
    first_row: tuple[object, ...] | None = next(row_iter, None)
    if first_row is None:
        return False
    cells_text: str = " ".join(
        str(c).lower() if c is not None else "" for c in first_row
    )
    return "topo" in cells_text or "lésion" in cells_text or "lesion" in cells_text


def extract_entries_standard(
    worksheet: "Worksheet", organe: Organe, feuille: str
) -> list[BiblesEntry]:
    """Extrait les entrees d'une feuille au schema standard [Topo, Lib, Ref, Texte].

    Tolere les feuilles sans colonne Topo : dans ce cas, ``topographie`` est
    vide et la lesion porte toute l'information. Tolere aussi les lignes
    vides et les textes incomplets.
    """
    entries: list[BiblesEntry] = []
    has_header: bool = detect_header_row(worksheet)
    start_row: int = 2 if has_header else 1

    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        values: list[str] = [normalize_text(c) for c in row]
        if not any(values):
            continue

        topo, lesion, code, texte = _unpack_row_cells(values)
        if len(texte) < MIN_TEXTE_LENGTH:
            continue

        entries.append(
            BiblesEntry(
                organe=organe,
                topographie=topo,
                lesion=lesion,
                code_adicap=code,
                texte_standard=texte,
                feuille_source=feuille,
            )
        )

    return entries


def _unpack_row_cells(values: list[str]) -> tuple[str, str, str, str]:
    """Deduit (topo, lesion, code, texte) d'une ligne selon sa longueur.

    Heuristique : une feuille a 4 colonnes = Topo+Lib+Ref+Texte,
    a 3 colonnes = Lib+Ref+Texte (pas de Topo), a 2 = Lib+Texte.
    La derniere cellule non vide est toujours consideree comme le texte
    standard (c'est la plus longue en pratique).
    """
    trimmed: list[str] = list(values)
    while trimmed and not trimmed[-1]:
        trimmed.pop()

    if len(trimmed) >= 4:
        return (trimmed[0], trimmed[1], trimmed[2], trimmed[3])
    if len(trimmed) == 3:
        return ("", trimmed[0], trimmed[1], trimmed[2])
    if len(trimmed) == 2:
        return ("", trimmed[0], "", trimmed[1])
    if len(trimmed) == 1:
        return ("", "", "", trimmed[0])
    return ("", "", "", "")


# ---------------------------------------------------------------------------
# Extracteurs speciaux pour les feuilles atypiques
# ---------------------------------------------------------------------------


def extract_entries_bdig(worksheet: "Worksheet") -> list[BiblesEntry]:
    """Feuille BDig : macro-templates digestifs en schema (nom, prose)."""
    entries: list[BiblesEntry] = []
    for row in worksheet.iter_rows(values_only=True):
        nom: str = normalize_text(row[0] if len(row) > 0 else None)
        texte: str = normalize_text(row[1] if len(row) > 1 else None)
        if len(texte) < MIN_TEXTE_LENGTH:
            continue
        entries.append(
            BiblesEntry(
                organe="digestif",
                topographie=nom,
                lesion="template_macroscopie",
                code_adicap="",
                texte_standard=texte,
                feuille_source="BDig",
            )
        )
    return entries


def extract_entries_single_col(worksheet: "Worksheet", feuille: str) -> list[BiblesEntry]:
    """Feuille CR (ou similaire) : une seule colonne de templates IHC generiques."""
    entries: list[BiblesEntry] = []
    for idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        texte: str = normalize_text(row[0] if len(row) > 0 else None)
        if len(texte) < MIN_TEXTE_LENGTH:
            continue
        entries.append(
            BiblesEntry(
                organe="generic",
                topographie="",
                lesion=f"template_{feuille.lower()}_{idx}",
                code_adicap="",
                texte_standard=texte,
                feuille_source=feuille,
            )
        )
    return entries


# ---------------------------------------------------------------------------
# Orchestration principale
# ---------------------------------------------------------------------------


def ingest_bibles_greg(xlsx_path: Path) -> list[BiblesEntry]:
    """Parse le classeur Excel complet et retourne toutes les entrees typees."""
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    all_entries: list[BiblesEntry] = []

    for sheet_name in workbook.sheetnames:
        if sheet_name in SHEETS_TO_SKIP:
            continue

        worksheet = workbook[sheet_name]

        if sheet_name == "BDig":
            all_entries.extend(extract_entries_bdig(worksheet))
            continue

        if sheet_name == "CR":
            all_entries.extend(extract_entries_single_col(worksheet, "CR"))
            continue

        organe: Organe | None = SHEET_TO_ORGANE.get(sheet_name)
        if organe is None:
            continue

        all_entries.extend(extract_entries_standard(worksheet, organe, sheet_name))

    workbook.close()
    return all_entries


def write_entries(entries: list[BiblesEntry], output_path: Path) -> None:
    """Serialise la liste d'entrees en JSON stable (tri + indent pour diff git)."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    payload: list[dict[str, str]] = [
        entry.model_dump() for entry in entries
    ]
    with output_path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, ensure_ascii=False, indent=2)


def summarize_entries(entries: list[BiblesEntry]) -> str:
    """Resume par organe pour le log : 'poumon: 12, sein: 19, ...'."""
    counts: dict[str, int] = {}
    for entry in entries:
        counts[entry.organe] = counts.get(entry.organe, 0) + 1
    parts: list[str] = [f"{organe}: {count}" for organe, count in sorted(counts.items())]
    return ", ".join(parts)


def main() -> None:
    """Point d'entree CLI : ingest → summary → write."""
    xlsx_path: Path = (
        REPO_ROOT
        / "Ressources"
        / "1. Codes Bible"
        / "Bibles Greg.xlsx"
    )
    output_path: Path = (
        REPO_ROOT / "backend" / "retrieval" / "data" / "bibles_greg.json"
    )

    if not xlsx_path.exists():
        raise FileNotFoundError(f"Source introuvable : {xlsx_path}")

    entries: list[BiblesEntry] = ingest_bibles_greg(xlsx_path)
    write_entries(entries, output_path)

    print(f"[ingest_bibles_greg] {len(entries)} entrees")
    print(f"[ingest_bibles_greg] {summarize_entries(entries)}")
    print(f"[ingest_bibles_greg] ecrit : {output_path}")


if __name__ == "__main__":
    main()
