"""Precalcul de l'index du thesaurus ADICAP officiel : XLSX -> JSON compact.

Le runtime (backend/adicap_ref.py) ne doit ni embarquer openpyxl ni ouvrir un
classeur Excel au demarrage : ce script fait le travail une fois pour toutes et
depose un JSON que le backend se contente de charger.

Deux particularites du thesaurus dictent le format de sortie :
- la colonne `code` n'est PAS un identifiant : 9683 lignes pour 9394 codes
  distincts, et meme le couple (dictionnaire, code) se repete. Le seul
  identifiant fiable est l'URI, qui sert donc de cle unique ;
- 201 concepts portent une `endDate` : ils sont obsoletes et ne doivent plus
  etre proposes, mais restent lisibles pour afficher un compte rendu ancien.
  On conserve donc la date de fin plutot que de filtrer a la construction.

Le JSON est stocke en lignes positionnelles (pas de dictionnaire par concept)
pour eviter de repeter 9683 fois les noms de champs ; l'ordre des colonnes est
publie dans l'en-tete du fichier, et les URI y sont raccourcies de leur prefixe
commun. Le libelle normalise n'est PAS stocke : il est recalcule au chargement
par text_utils, source unique de normalisation du depot.

Usage :
  python scripts/build_adicap_index.py [chemin/vers/ADICAP_2024.xlsx]
"""

from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import TypedDict

import openpyxl  # type: ignore[import-untyped]

_RACINE_DEPOT: Path = Path(__file__).resolve().parent.parent

_XLSX_DEFAUT: Path = (
    _RACINE_DEPOT
    / "docs/specs/referentiels/terminologie-adicap-2024-10/dat/ADICAP_2024.xlsx"
)
_DESTINATION: Path = _RACINE_DEPOT / "backend/data/adicap_ref.json"

_FEUILLE: str = "adicapV3"
_VERSION: str = "ADICAP 2024-10"
_BASE_URI: str = "https://data.esante.gouv.fr/adicap/"

# Ordre des champs dans chaque ligne de `concepts`. Publie dans le JSON pour que
# le lecteur puisse refuser un fichier dont le schema aurait bouge.
_COLONNES: tuple[str, ...] = (
    "uri",
    "code",
    "uri_parent",
    "dictionnaire",
    "libelle",
    "fin_validite",
    "code_anatomie",
)


class IndexAdicapJSON(TypedDict):
    """Schema du fichier produit."""

    version: str
    source: str
    base_uri: str
    colonnes: list[str]
    anatomies: dict[str, str]
    concepts: list[list[str]]


def _lire_lignes(xlsx: Path) -> list[dict[str, str]]:
    """Lit la feuille adicapV3 et renvoie une ligne par concept, cle = en-tete."""
    classeur = openpyxl.load_workbook(str(xlsx), read_only=True, data_only=True)
    if _FEUILLE not in classeur.sheetnames:
        raise ValueError(f"feuille {_FEUILLE} absente de {xlsx}")
    feuille = classeur[_FEUILLE]
    lignes = feuille.iter_rows(values_only=True)
    entetes = [str(cellule).strip() for cellule in next(lignes)]
    return [
        {
            entete: "" if valeur is None else str(valeur).strip()
            for entete, valeur in zip(entetes, cellules)
        }
        for cellules in lignes
    ]


def _suffixe_uri(uri: str) -> str:
    """Retire le prefixe commun d'une URI ADICAP ('.../D1H' -> 'D1H').

    Une URI hors prefixe signalerait un fichier source inattendu : on refuse
    plutot que de produire un index silencieusement incoherent.
    """
    if not uri:
        return ""
    if not uri.startswith(_BASE_URI):
        raise ValueError(f"URI hors referentiel ADICAP : {uri}")
    return uri[len(_BASE_URI):]


def _jour(horodatage: str) -> str:
    """'1999-10-06T00:00:00Z' -> '1999-10-06' ; l'heure ne porte aucun sens ici."""
    return horodatage[:10]


def _concept(ligne: dict[str, str]) -> list[str]:
    """Convertit une ligne du classeur en ligne positionnelle de l'index."""
    return [
        _suffixe_uri(ligne["URI"]),
        ligne["code"],
        _suffixe_uri(ligne["Parent"]),
        ligne["dictionaryCode"],
        ligne["libelle"],
        _jour(ligne["endDate"]),
        ligne["anatomyCode"],
    ]


def _anatomies(lignes: list[dict[str, str]]) -> dict[str, str]:
    """Table code anatomique -> libelle, sortie du corps pour ne pas la repeter.

    126 codes anatomiques seulement pour plus de 4000 concepts porteurs : les
    stocker a part evite de recopier le libelle sur chaque ligne.
    """
    table: dict[str, str] = {}
    for ligne in lignes:
        code = ligne["anatomyCode"]
        if code:
            table[code] = ligne["anatomyLabel"]
    return dict(sorted(table.items()))


def construire_index(xlsx: Path) -> IndexAdicapJSON:
    """Assemble l'index complet a partir du classeur officiel."""
    lignes = _lire_lignes(xlsx)
    return {
        "version": _VERSION,
        "source": xlsx.name,
        "base_uri": _BASE_URI,
        "colonnes": list(_COLONNES),
        "anatomies": _anatomies(lignes),
        "concepts": [_concept(ligne) for ligne in lignes],
    }


def _verifier(index: IndexAdicapJSON) -> None:
    """Refuse d'ecrire un index dont l'arbre serait casse.

    Un index publie avec des URI dupliquees ou des parents manquants ferait
    echouer le runtime bien plus tard et bien plus obscurement.
    """
    concepts = index["concepts"]
    uris = [ligne[0] for ligne in concepts]
    if len(set(uris)) != len(uris):
        raise ValueError("URI dupliquees : l'URI doit etre la cle unique")
    connues = set(uris)
    orphelins = [
        ligne[0] for ligne in concepts if ligne[2] and ligne[2] not in connues
    ]
    if orphelins:
        raise ValueError(f"{len(orphelins)} concepts orphelins, ex. {orphelins[:3]}")
    racines = [ligne[0] for ligne in concepts if not ligne[2]]
    if len(racines) != 1:
        raise ValueError(f"arbre a {len(racines)} racines, une seule attendue")


def _ecrire(index: IndexAdicapJSON, destination: Path) -> int:
    """Ecrit le JSON (un concept par ligne) et renvoie sa taille en octets.

    Format manuel plutot que json.dump : compact sur les separateurs, mais une
    ligne par concept pour rester lisible et diffable dans git.
    """
    destination.parent.mkdir(parents=True, exist_ok=True)
    entete = [
        f' {json.dumps(cle)}: {json.dumps(valeur, ensure_ascii=False)},'
        for cle, valeur in index.items()
        if cle != "concepts"
    ]
    concepts = [
        "  " + json.dumps(ligne, ensure_ascii=False, separators=(",", ":"))
        for ligne in index["concepts"]
    ]
    texte = (
        "{\n" + "\n".join(entete) + '\n "concepts": [\n'
        + ",\n".join(concepts) + "\n ]\n}\n"
    )
    destination.write_text(texte, encoding="utf-8")
    return len(texte.encode("utf-8"))


def _profondeur_max(index: IndexAdicapJSON) -> int:
    """Profondeur du concept le plus enfoui, indicateur de sante de l'arbre."""
    parents = {ligne[0]: ligne[2] for ligne in index["concepts"]}
    profondeurs: dict[str, int] = {}

    def profondeur(uri: str) -> int:
        if uri not in profondeurs:
            parent = parents[uri]
            profondeurs[uri] = 0 if not parent else profondeur(parent) + 1
        return profondeurs[uri]

    return max(profondeur(uri) for uri in parents)


def _rapporter(index: IndexAdicapJSON, octets: int) -> None:
    """Affiche les effectifs mesures pour controle a chaque reconstruction."""
    concepts = index["concepts"]
    effectifs: dict[str, int] = {}
    for ligne in concepts:
        effectifs[ligne[3]] = effectifs.get(ligne[3], 0) + 1
    obsoletes = sum(1 for ligne in concepts if ligne[5])
    print(f"{len(concepts)} concepts -> {_DESTINATION}")
    print(f"  taille        : {octets} octets ({octets / 1024:.0f} Kio)")
    print(f"  dictionnaires : {dict(sorted(effectifs.items()))}")
    print(f"  obsoletes     : {obsoletes}")
    print(f"  anatomies     : {len(index['anatomies'])}")
    print(f"  profondeur max: {_profondeur_max(index)}")


def main() -> int:
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else _XLSX_DEFAUT
    if not xlsx.exists():
        print(f"Classeur introuvable : {xlsx}", file=sys.stderr)
        return 1
    index = construire_index(xlsx)
    _verifier(index)
    _rapporter(index, _ecrire(index, _DESTINATION))
    return 0


if __name__ == "__main__":
    sys.exit(main())
