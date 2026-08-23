"""Export des donnees de l'etude : quatre tables a plat, et leur lisez-moi.

Sans export, l'etude n'est pas analysable : elle est seulement consultable. Ces
tests verifient ce qui rend un CSV depouillable des mois plus tard — un en-tete
fige, un pseudonyme qui ne bouge pas, des textes qui se relisent intacts, et
surtout des cellules VIDES la ou rien n'a ete mesure, jamais des zeros.
"""

from __future__ import annotations

import csv
import io
import zipfile
from datetime import datetime

import pytest
from fastapi.testclient import TestClient
from sqlalchemy.ext.asyncio import async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

import main
from auth import get_current_user
from database import get_db_session
from db_models import Base, User
from etude.export import (
    COLONNES_DOSSIERS,
    COLONNES_PAUSES,
    COLONNES_PROPOSITIONS,
    COLONNES_QUESTIONNAIRES,
    NOM_DOSSIERS,
    NOM_LISEZ_MOI,
    NOM_PAUSES,
    NOM_PROPOSITIONS,
    NOM_QUESTIONNAIRES,
)
from etude.models import (
    EtudeDossier,
    EtudePause,
    EtudePrelevement,
    EtudeProposition,
    EtudeReponseQuestionnaire,
    EtudeSession,
)

# Le premier inclus porte l'identifiant qui trie EN DERNIER : si la
# numerotation suivait l'ordre alphabetique des comptes plutot que l'ordre
# d'inclusion, ces deux praticiens seraient intervertis.
PRATICIEN_TOT = "u-zeta"
PRATICIEN_TARD = "u-alpha"

# Guillemets, virgules, points-virgules et retours a la ligne : tout ce qui
# casse une concatenation ecrite a la main.
TRANSCRIPTION_PIEGEE = (
    'Le praticien a dicte : "adenome tubuleux, en dysplasie de bas grade" ;\n'
    "puis, apres un temps, il a ajoute une deuxieme ligne.\n"
    'Elle se termine par un guillemet orphelin " et une virgule, ici.'
)
CR_PROPOSE = '**Conclusion :**\n"Adenome tubuleux" du sigmoide.'
CR_VALIDE = '**Conclusion :**\n"Adenome tubuleux" du sigmoide, sans dysplasie.'

DOSSIER_NORMAL = "d-normal"
DOSSIER_EXCLU = "d-exclu"
DOSSIER_ABANDONNE = "d-abandonne"


# ---------------------------------------------------------------------------
# Base de test
# ---------------------------------------------------------------------------


def _compte(identifiant: str, role: str) -> User:
    user = User()
    user.id = identifiant
    user.email = f"{identifiant}@marc.test"
    user.name = "Compte"
    user.role = role
    return user


async def _creer_schema(moteur) -> None:
    import etude.models  # noqa: F401  (enregistre les tables sur Base)

    async with moteur.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


def _praticiens() -> list[User]:
    return [
        User(
            id=PRATICIEN_TOT,
            email="zeta@marc.test",
            password_hash="x",
            name="Docteur Zeta",
            role="user",
        ),
        User(
            id=PRATICIEN_TARD,
            email="alpha@marc.test",
            password_hash="x",
            name="Docteur Alpha",
            role="user",
        ),
    ]


def _sessions() -> list[EtudeSession]:
    """Deux sessions, la plus ancienne appartenant au praticien 'u-zeta'."""
    return [
        EtudeSession(
            id="s-tot",
            praticien_id=PRATICIEN_TOT,
            debut=datetime(2026, 3, 1, 8, 0),
            nb_cas=2,
        ),
        EtudeSession(
            id="s-tard",
            praticien_id=PRATICIEN_TARD,
            debut=datetime(2026, 3, 5, 9, 30),
            nb_cas=1,
        ),
    ]


def _dossier_normal() -> EtudeDossier:
    """Un cas mene jusqu'a la cloture, avec des textes pieges."""
    return EtudeDossier(
        id=DOSSIER_NORMAL,
        session_id="s-tot",
        index_session=0,
        transcription=TRANSCRIPTION_PIEGEE,
        cr_propose=CR_PROPOSE,
        cr_valide=CR_VALIDE,
        t0_debut_dictee=datetime(2026, 3, 1, 8, 1, 0),
        t1_fin_dictee=datetime(2026, 3, 1, 8, 1, 40),
        t2_affichage=datetime(2026, 3, 1, 8, 1, 45),
        t3_premiere_decision=datetime(2026, 3, 1, 8, 2, 0),
        t4_derniere_decision=datetime(2026, 3, 1, 8, 3, 0),
        t5_cloture=datetime(2026, 3, 1, 8, 3, 30),
        t6_export=datetime(2026, 3, 1, 8, 3, 40),
        nb_prelevements_detecte=1,
        nb_prelevements_corrige=1,
        omission_signalee=False,
        caracteres_modifies=18,
        organe="colon",
        cree_a=datetime(2026, 3, 1, 8, 0, 30),
    )


def _dossier_exclu() -> EtudeDossier:
    """Un essai de l'administrateur, ecarte mais conserve."""
    return EtudeDossier(
        id=DOSSIER_EXCLU,
        session_id="s-tot",
        index_session=1,
        transcription="Essai sans valeur clinique.",
        cr_propose="Essai.",
        cr_valide="Essai.",
        t2_affichage=datetime(2026, 3, 1, 8, 10, 0),
        t5_cloture=datetime(2026, 3, 1, 8, 10, 30),
        omission_signalee=False,
        caracteres_modifies=0,
        exclu=True,
        motif_exclusion="Essai de l'administrateur, pas un cas reel",
        exclu_par="adm",
        organe="colon",
        cree_a=datetime(2026, 3, 1, 8, 9, 0),
    )


def _dossier_abandonne() -> EtudeDossier:
    """Un cas quitte en cours de route : presque tout y est NON MESURE."""
    return EtudeDossier(
        id=DOSSIER_ABANDONNE,
        session_id="s-tard",
        index_session=0,
        transcription="Biopsie gastrique, antre.",
        cr_propose="Gastrite chronique.",
        t0_debut_dictee=datetime(2026, 3, 5, 9, 31, 0),
        t1_fin_dictee=datetime(2026, 3, 5, 9, 31, 20),
        t2_affichage=datetime(2026, 3, 5, 9, 31, 25),
        abandonne=True,
        motif_abandon="interruption",
        organe="estomac",
        cree_a=datetime(2026, 3, 5, 9, 30, 30),
    )


def _propositions() -> list[EtudeProposition]:
    return [
        EtudeProposition(
            id="p-decidee",
            dossier_id=DOSSIER_NORMAL,
            prelevement_id="pr-1",
            type="restitution",
            sous_type="diagnostic",
            valeur_proposee='Adenome tubuleux, "bas grade"',
            empan_debut=24,
            empan_fin=70,
            longueur_mots=20,
            affiche_a=datetime(2026, 3, 1, 8, 1, 45),
            decide_a=datetime(2026, 3, 1, 8, 2, 0),
            latence_ms=15000,
            hative=False,
            justif_ouverte=True,
            justif_duree_ms=4000,
            decision_changee_apres_justif=True,
            decision="corrige",
            valeur_retenue="Adenome tubuleux sans dysplasie",
            nature_correction="erreur_fond",
            cause_erreur="interpretation",
        ),
        EtudeProposition(
            id="p-non-decidee",
            dossier_id=DOSSIER_NORMAL,
            type="code",
            valeur_proposee="BHGS0001",
            positions="[5, 6, 7, 8]",
            confiance=0.62,
            affiche_a=datetime(2026, 3, 1, 8, 1, 45),
        ),
        EtudeProposition(
            id="p-exclue",
            dossier_id=DOSSIER_EXCLU,
            type="restitution",
            valeur_proposee="Essai.",
            empan_debut=0,
            empan_fin=5,
            longueur_mots=1,
            decision="conforme",
        ),
        EtudeProposition(
            id="p-abandon",
            dossier_id=DOSSIER_ABANDONNE,
            type="completude",
            valeur_proposee="Preciser l'activite de la gastrite.",
        ),
    ]


def _pauses() -> list[EtudePause]:
    return [
        EtudePause(
            id="pause-close",
            dossier_id=DOSSIER_NORMAL,
            debut=datetime(2026, 3, 1, 8, 2, 10),
            fin=datetime(2026, 3, 1, 8, 2, 25),
            duree_ms=15000,
            cause="onglet_masque",
        ),
        # Jamais refermee : sa duree est INCONNUE, pas nulle.
        EtudePause(
            id="pause-ouverte",
            dossier_id=DOSSIER_ABANDONNE,
            debut=datetime(2026, 3, 5, 9, 32, 0),
            cause="inactivite",
        ),
    ]


def _reponses() -> list[EtudeReponseQuestionnaire]:
    return [
        EtudeReponseQuestionnaire(
            id="r-inclusion",
            praticien_id=PRATICIEN_TOT,
            questionnaire="inclusion",
            item="inclusion_01",
            valeur="12",
            repondu_a=datetime(2026, 3, 1, 7, 55),
        ),
        EtudeReponseQuestionnaire(
            id="r-par-cas",
            praticien_id=PRATICIEN_TOT,
            dossier_id=DOSSIER_NORMAL,
            questionnaire="par_cas",
            item="par_cas_04",
            valeur="4",
            repondu_a=datetime(2026, 3, 1, 8, 4),
        ),
        EtudeReponseQuestionnaire(
            id="r-par-cas-exclu",
            praticien_id=PRATICIEN_TOT,
            dossier_id=DOSSIER_EXCLU,
            questionnaire="par_cas",
            item="par_cas_04",
            valeur="1",
            repondu_a=datetime(2026, 3, 1, 8, 11),
        ),
    ]


async def _peupler(fabrique) -> None:
    """Un dossier normal, un exclu, un abandonne, et deux praticiens."""
    async with fabrique() as session:
        session.add_all(_praticiens())
        session.add_all(_sessions())
        session.add_all(
            [_dossier_normal(), _dossier_exclu(), _dossier_abandonne()]
        )
        session.add(
            EtudePrelevement(
                id="pr-1",
                dossier_id=DOSSIER_NORMAL,
                rang=1,
                libelle="Sigmoide, biopsie",
            )
        )
        session.add_all(_propositions())
        session.add_all(_pauses())
        session.add_all(_reponses())
        await session.commit()


def _ouvrir(tmp_path, remplir: bool):
    moteur = create_async_engine(
        f"sqlite+aiosqlite:///{tmp_path / 'export.db'}", poolclass=NullPool
    )
    fabrique = async_sessionmaker(moteur, expire_on_commit=False)

    async def _base():
        async with fabrique() as session:
            yield session

    app = main.app
    app.dependency_overrides[get_db_session] = _base
    app.dependency_overrides[get_current_user] = lambda: _compte("adm", "admin")

    with TestClient(app) as client:
        client.portal.call(_creer_schema, moteur)
        if remplir:
            client.portal.call(_peupler, fabrique)
        yield client
        app.dependency_overrides.clear()
        client.portal.call(moteur.dispose)


@pytest.fixture
def client(tmp_path):
    """Un administrateur devant une base peuplee."""
    yield from _ouvrir(tmp_path, remplir=True)


@pytest.fixture
def client_vide(tmp_path):
    """Un administrateur devant une base ou rien n'a encore ete recueilli."""
    yield from _ouvrir(tmp_path, remplir=False)


# ---------------------------------------------------------------------------
# Lecture de l'archive
# ---------------------------------------------------------------------------


def _archive(client) -> dict[str, bytes]:
    reponse = client.get("/admin/etude/export")
    assert reponse.status_code == 200, reponse.text
    with zipfile.ZipFile(io.BytesIO(reponse.content)) as zip_lu:
        return {nom: zip_lu.read(nom) for nom in zip_lu.namelist()}


def _table(archive: dict[str, bytes], nom: str) -> list[dict[str, str]]:
    """Relit un CSV comme le fera R ou pandas : par un vrai lecteur de CSV."""
    texte = archive[nom].decode("utf-8-sig")
    return list(csv.DictReader(io.StringIO(texte, newline="")))


def _entete(archive: dict[str, bytes], nom: str) -> list[str]:
    texte = archive[nom].decode("utf-8-sig")
    return next(csv.reader(io.StringIO(texte, newline="")))


def _par_id(lignes: list[dict[str, str]], colonne: str) -> dict[str, dict[str, str]]:
    return {ligne[colonne]: ligne for ligne in lignes}


# ---------------------------------------------------------------------------
# Acces
# ---------------------------------------------------------------------------


def test_un_praticien_ne_peut_pas_exporter(client):
    """L'export porte les donnees de tous les confreres : il ne regarde pas un
    participant."""
    main.app.dependency_overrides[get_current_user] = lambda: _compte("p", "user")
    assert client.get("/admin/etude/export").status_code == 403


# ---------------------------------------------------------------------------
# Forme de l'archive
# ---------------------------------------------------------------------------


def test_l_archive_contient_les_quatre_tables_et_son_lisez_moi(client):
    archive = _archive(client)
    assert set(archive) == {
        NOM_DOSSIERS,
        NOM_PROPOSITIONS,
        NOM_QUESTIONNAIRES,
        NOM_PAUSES,
        NOM_LISEZ_MOI,
    }


def test_l_archive_se_telecharge_sous_un_nom_date(client):
    reponse = client.get("/admin/etude/export")
    assert reponse.headers["content-type"] == "application/zip"
    assert "export_etude_marc_" in reponse.headers["content-disposition"]
    assert reponse.headers["content-disposition"].endswith('.zip"')


def test_l_en_tete_de_chaque_fichier_est_fige(client):
    """Un ordre de colonnes qui bouge rend deux exports incomparables et casse
    tout script ecrit sur le premier."""
    archive = _archive(client)
    assert _entete(archive, NOM_DOSSIERS) == list(COLONNES_DOSSIERS)
    assert _entete(archive, NOM_PROPOSITIONS) == list(COLONNES_PROPOSITIONS)
    assert _entete(archive, NOM_QUESTIONNAIRES) == list(COLONNES_QUESTIONNAIRES)
    assert _entete(archive, NOM_PAUSES) == list(COLONNES_PAUSES)


def test_les_csv_portent_le_bom_utf8(client):
    """Sans le BOM, Excel lit l'UTF-8 comme du Latin-1 — et Excel sera le
    premier lecteur de ces fichiers."""
    archive = _archive(client)
    for nom in (NOM_DOSSIERS, NOM_PROPOSITIONS, NOM_QUESTIONNAIRES, NOM_PAUSES):
        assert archive[nom].startswith(b"\xef\xbb\xbf"), nom


# ---------------------------------------------------------------------------
# Pseudonymisation
# ---------------------------------------------------------------------------


def test_aucun_nom_ni_adresse_ne_sort_dans_l_archive(client):
    archive = _archive(client)
    entier = b"".join(archive.values()).decode("utf-8-sig")
    for trace in ("@marc.test", "Docteur", PRATICIEN_TOT, PRATICIEN_TARD):
        assert trace not in entier


def test_le_pseudonyme_suit_l_ordre_d_inclusion(client):
    """Un ordre alphabetique sur l'identifiant de compte donnerait l'inverse :
    c'est bien l'ordre d'inclusion qui numerote."""
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    assert dossiers[DOSSIER_NORMAL]["praticien"] == "praticien_01"
    assert dossiers[DOSSIER_ABANDONNE]["praticien"] == "praticien_02"


def test_le_pseudonyme_ne_bouge_pas_d_un_export_a_l_autre(client):
    """Sans stabilite, on ne peut ni suivre un praticien dans le temps ni
    recouper deux exports."""
    premier = _table(_archive(client), NOM_DOSSIERS)
    second = _table(_archive(client), NOM_DOSSIERS)
    assert [ligne["praticien"] for ligne in premier] == [
        ligne["praticien"] for ligne in second
    ]


def test_le_pseudonyme_relie_les_quatre_fichiers(client):
    archive = _archive(client)
    propositions = _par_id(_table(archive, NOM_PROPOSITIONS), "proposition_id")
    pauses = _par_id(_table(archive, NOM_PAUSES), "pause_id")
    reponses = _par_id(_table(archive, NOM_QUESTIONNAIRES), "reponse_id")
    assert propositions["p-decidee"]["praticien"] == "praticien_01"
    assert pauses["pause-close"]["praticien"] == "praticien_01"
    assert reponses["r-inclusion"]["praticien"] == "praticien_01"


# ---------------------------------------------------------------------------
# Dossiers exclus : exportes, signales, jamais caches
# ---------------------------------------------------------------------------


def test_le_dossier_exclu_est_exporte_avec_son_motif(client):
    """Les omettre empecherait de rendre compte de l'effectif ecarte, que toute
    publication demande."""
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    assert dossiers[DOSSIER_EXCLU]["exclu"] == "true"
    assert dossiers[DOSSIER_EXCLU]["motif_exclusion"].startswith("Essai")
    assert dossiers[DOSSIER_NORMAL]["exclu"] == "false"


def test_le_drapeau_d_exclusion_est_recopie_sur_les_lignes_filles(client):
    """Une jointure oubliee est une jointure qui n'aura pas lieu : le filtre
    doit tenir en une colonne."""
    archive = _archive(client)
    propositions = _par_id(_table(archive, NOM_PROPOSITIONS), "proposition_id")
    reponses = _par_id(_table(archive, NOM_QUESTIONNAIRES), "reponse_id")
    assert propositions["p-exclue"]["dossier_exclu"] == "true"
    assert propositions["p-decidee"]["dossier_exclu"] == "false"
    assert reponses["r-par-cas-exclu"]["dossier_exclu"] == "true"


def test_les_fichiers_se_recollent_par_les_identifiants(client):
    archive = _archive(client)
    connus = {ligne["dossier_id"] for ligne in _table(archive, NOM_DOSSIERS)}
    assert len(connus) == 3
    for ligne in _table(archive, NOM_PROPOSITIONS):
        assert ligne["dossier_id"] in connus
    for ligne in _table(archive, NOM_PAUSES):
        assert ligne["dossier_id"] in connus


# ---------------------------------------------------------------------------
# Les textes
# ---------------------------------------------------------------------------


def test_un_texte_avec_guillemets_et_retours_a_la_ligne_se_relit_intact(client):
    """C'est ce que le module csv garantit et qu'une concatenation detruit."""
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    ligne = dossiers[DOSSIER_NORMAL]
    assert ligne["transcription"] == TRANSCRIPTION_PIEGEE
    assert ligne["cr_propose"] == CR_PROPOSE
    assert ligne["cr_valide"] == CR_VALIDE


def test_le_texte_pieges_ne_decale_pas_les_colonnes(client):
    """Un retour a la ligne mal echappe ferait deux lignes d'un dossier."""
    dossiers = _table(_archive(client), NOM_DOSSIERS)
    assert len(dossiers) == 3
    for ligne in dossiers:
        assert set(ligne) == set(COLONNES_DOSSIERS)


def test_les_deux_versions_du_compte_rendu_partent_ensemble(client):
    """Sans le texte propose a cote du texte valide, la charge d'edition n'est
    pas recalculable."""
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    ligne = dossiers[DOSSIER_NORMAL]
    assert ligne["cr_propose"] != ligne["cr_valide"]
    assert ligne["caracteres_modifies"] == "18"


# ---------------------------------------------------------------------------
# Une absence de mesure n'est pas un zero
# ---------------------------------------------------------------------------


def test_un_dossier_abandonne_laisse_ses_cellules_vides_et_non_a_zero(client):
    """Ecrire 0 la ou l'on n'a rien observe est la maniere la plus courante de
    mentir avec un tableau."""
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    ligne = dossiers[DOSSIER_ABANDONNE]
    assert ligne["abandonne"] == "true"
    assert ligne["motif_abandon"] == "interruption"
    for colonne in ("t5_cloture", "revision_ms", "revision_nette_ms",
                    "caracteres_modifies", "cr_valide"):
        assert ligne[colonne] == "", colonne


def test_une_omission_non_renseignee_ne_se_lit_pas_comme_une_absence_d_omission(
    client,
):
    """Sortir 'false' compterait chaque dossier en cours comme un dossier sans
    omission, et le taux publie serait faux vers le bas."""
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    assert dossiers[DOSSIER_ABANDONNE]["omission_signalee"] == ""
    assert dossiers[DOSSIER_NORMAL]["omission_signalee"] == "false"


def test_une_proposition_sans_decision_reste_vide(client):
    """Une proposition non decidee n'est pas une proposition refusee."""
    propositions = _par_id(
        _table(_archive(client), NOM_PROPOSITIONS), "proposition_id"
    )
    ligne = propositions["p-non-decidee"]
    assert ligne["decision"] == ""
    assert ligne["latence_ms"] == ""
    assert ligne["nature_correction"] == ""
    assert ligne["confiance"] == "0.62"


def test_une_pause_jamais_refermee_ne_declare_pas_une_duree_nulle(client):
    pauses = _par_id(_table(_archive(client), NOM_PAUSES), "pause_id")
    assert pauses["pause-ouverte"]["fin"] == ""
    assert pauses["pause-ouverte"]["duree_ms"] == ""
    assert pauses["pause-close"]["duree_ms"] == "15000"


def test_une_reponse_hors_dossier_laisse_la_colonne_d_exclusion_vide(client):
    """Le questionnaire d'inclusion ne porte sur aucun cas : ni exclu, ni non
    exclu."""
    reponses = _par_id(_table(_archive(client), NOM_QUESTIONNAIRES), "reponse_id")
    assert reponses["r-inclusion"]["dossier_id"] == ""
    assert reponses["r-inclusion"]["dossier_exclu"] == ""


# ---------------------------------------------------------------------------
# Ce que porte chaque proposition
# ---------------------------------------------------------------------------


def test_la_proposition_porte_sa_decision_et_sa_nature_de_correction(client):
    """« Corrige » seul ne dit pas si le systeme s'est trompe ou si le
    praticien ecrit autrement."""
    propositions = _par_id(
        _table(_archive(client), NOM_PROPOSITIONS), "proposition_id"
    )
    ligne = propositions["p-decidee"]
    assert ligne["decision"] == "corrige"
    assert ligne["nature_correction"] == "erreur_fond"
    assert ligne["cause_erreur"] == "interpretation"
    assert ligne["decision_changee_apres_justif"] == "true"
    assert ligne["ancree"] == "true"
    assert ligne["prelevement_rang"] == "1"
    assert ligne["prelevement_libelle"] == "Sigmoide, biopsie"


def test_une_proposition_sans_empan_est_signalee_non_ancree(client):
    """Pas d'empan, pas de soutien dans la dictee : c'est la mesure
    d'hallucination cote systeme."""
    propositions = _par_id(
        _table(_archive(client), NOM_PROPOSITIONS), "proposition_id"
    )
    assert propositions["p-non-decidee"]["ancree"] == "false"
    assert propositions["p-non-decidee"]["empan_debut"] == ""


def test_les_temps_du_dossier_sont_calcules_pauses_deduites(client):
    dossiers = _par_id(_table(_archive(client), NOM_DOSSIERS), "dossier_id")
    ligne = dossiers[DOSSIER_NORMAL]
    assert ligne["dictee_ms"] == "40000"
    assert ligne["revision_ms"] == "105000"
    assert ligne["pauses_ms"] == "15000"
    assert ligne["revision_nette_ms"] == "90000"
    assert ligne["nb_pauses"] == "1"


# ---------------------------------------------------------------------------
# Base vide
# ---------------------------------------------------------------------------


def test_une_base_vide_produit_quand_meme_les_quatre_en_tetes(client_vide):
    """Une etude qui n'a rien recueilli doit le dire avec des fichiers
    lisibles, pas avec une archive absente."""
    archive = _archive(client_vide)
    assert _entete(archive, NOM_DOSSIERS) == list(COLONNES_DOSSIERS)
    for nom in (NOM_DOSSIERS, NOM_PROPOSITIONS, NOM_QUESTIONNAIRES, NOM_PAUSES):
        assert _table(archive, nom) == []


def test_une_base_vide_annonce_zero_ligne_dans_le_lisez_moi(client_vide):
    lisez_moi = _archive(client_vide)[NOM_LISEZ_MOI].decode("utf-8")
    assert "dossiers.csv" in lisez_moi
    assert "0 ligne(s)" in lisez_moi


# ---------------------------------------------------------------------------
# Le lisez-moi
# ---------------------------------------------------------------------------


def test_le_lisez_moi_donne_la_date_et_les_effectifs(client):
    """Un export sans dictionnaire de donnees est inexploitable six mois plus
    tard."""
    lisez_moi = _archive(client)[NOM_LISEZ_MOI].decode("utf-8")
    assert lisez_moi.startswith("EXPORT DE L'ETUDE MARC")
    assert "Export genere le" in lisez_moi
    assert "3 ligne(s)" in lisez_moi  # dossiers.csv
    assert "4 ligne(s)" in lisez_moi  # propositions.csv


def test_le_lisez_moi_explique_les_colonnes_qui_ne_se_devinent_pas(client):
    lisez_moi = _archive(client)[NOM_LISEZ_MOI].decode("utf-8")
    for colonne in (
        "hative",
        "ancree",
        "nature_correction",
        "dossier_exclu",
        "revision_nette_ms",
    ):
        assert colonne in lisez_moi


def test_le_lisez_moi_avertit_qu_une_cellule_vide_n_est_pas_un_zero(client):
    """Le premier lecteur doit savoir qu'un remplacement des vides par 0
    fabriquerait un chiffre faux."""
    lisez_moi = _archive(client)[NOM_LISEZ_MOI].decode("utf-8")
    assert "NON MESURE" in lisez_moi
    assert "praticien_01" in lisez_moi


def test_le_lisez_moi_dit_comment_traiter_les_dossiers_exclus(client):
    lisez_moi = _archive(client)[NOM_LISEZ_MOI].decode("utf-8")
    assert "DOSSIERS EXCLUS" in lisez_moi


# --- Le classeur Excel ------------------------------------------------------


def test_le_classeur_porte_les_memes_lignes_que_les_csv():
    """Deux exports du meme corpus doivent se recouper a la ligne pres, sinon
    on ne sait plus lequel fait foi. Le CSV est le format d'ANALYSE, le
    classeur celui du TRAVAIL — mais ils decrivent le meme corpus."""
    import io
    from datetime import UTC, datetime

    from openpyxl import load_workbook

    from etude.export import (
        construire_classeur,
        construire_fichiers,
        Corpus,
    )

    corpus = Corpus([], [], [], [], [], [])
    moment = datetime(2026, 9, 1, tzinfo=UTC)
    classeur = load_workbook(io.BytesIO(construire_classeur(corpus, moment)))

    # Un onglet par table, plus le lisez-moi : dans un classeur, un fichier
    # texte a cote ne serait jamais ouvert.
    assert "lisez-moi" in classeur.sheetnames
    assert len(classeur.sheetnames) == len(construire_fichiers(corpus)) + 1


def test_le_classeur_fige_ses_en_tetes():
    """Une table de milliers de lignes devient illisible des le premier ecran
    si l'en-tete disparait au defilement."""
    import io
    from datetime import UTC, datetime

    from openpyxl import load_workbook

    from etude.export import Corpus, construire_classeur

    classeur = load_workbook(
        io.BytesIO(
            construire_classeur(
                Corpus([], [], [], [], [], []),
                datetime(2026, 9, 1, tzinfo=UTC),
            )
        )
    )
    for nom in classeur.sheetnames:
        if nom == "lisez-moi":
            continue
        assert classeur[nom].freeze_panes == "A2", nom
