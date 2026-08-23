"""Export des donnees de l'etude en tableaux a plat.

Aucun depouillement serieux ne se fait dans un navigateur. Sans export, l'etude
n'est pas analysable : elle est seulement consultable, ce qui n'est pas la meme
chose.

QUATRE TABLES A PLAT, une par fichier, reliees par des identifiants. A plat et
non imbrique, parce qu'un CSV s'ouvre dans R, dans Python et dans un tableur,
alors qu'un JSON imbrique impose d'ecrire un script avant de commencer — et
c'est exactement ce qui fait qu'on ne depouille jamais.

Trois regles gouvernent chaque cellule produite ici :

1. UNE CELLULE VIDE SIGNIFIE "NON MESURE". Jamais zero, jamais faux. Ecrire 0
   la ou l'on n'a rien observe est la maniere la plus courante de mentir avec
   un tableau, et le lecteur d'un CSV n'a aucun moyen de rattraper l'erreur.
2. LES DOSSIERS EXCLUS SONT EXPORTES, avec leur motif, et la colonne `exclu`
   est repetee sur les tables filles. Les omettre empecherait de rendre compte
   de l'effectif ecarte, que toute publication demande : c'est a l'analyse de
   filtrer, pas a l'export de cacher.
3. LE PRATICIEN EST PSEUDONYMISE DE FACON STABLE. Aucun nom, aucun email,
   aucun identifiant technique de compte ne sort d'ici.

Reference : docs/specs/etude/Protocole_etude_MARC.md.
"""

from __future__ import annotations

import csv
import io
import textwrap
import zipfile
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime, timezone
from typing import Final, TypeVar

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from etude.analyse import calculer_temps
from etude.models import (
    EtudeDossier,
    EtudePause,
    EtudePrelevement,
    EtudeProposition,
    EtudeReponseQuestionnaire,
    EtudeSession,
)

# --- Noms des fichiers de l'archive ----------------------------------------

NOM_DOSSIERS: Final = "dossiers.csv"
NOM_PROPOSITIONS: Final = "propositions.csv"
NOM_QUESTIONNAIRES: Final = "questionnaires.csv"
NOM_PAUSES: Final = "pauses.csv"
NOM_REVISIONS: Final = "revisions_decision.csv"
NOM_LISEZ_MOI: Final = "lisez_moi.txt"


# --- En-tetes ---------------------------------------------------------------
#
# Declares en constantes, et non deduits d'un parcours de dictionnaire : un
# ordre de colonnes qui bouge d'un export a l'autre rend les deux exports
# incomparables, et casse tout script d'analyse ecrit sur le premier.

COLONNES_DOSSIERS: Final[tuple[str, ...]] = (
    "dossier_id",
    "session_id",
    "praticien",
    "index_session",
    "cree_a",
    "organe",
    "exclu",
    "motif_exclusion",
    "abandonne",
    "motif_abandon",
    "omission_signalee",
    "omission_texte",
    "nb_prelevements_detecte",
    "nb_prelevements_corrige",
    "nb_propositions",
    "nb_decidees",
    "caracteres_modifies",
    "t0_debut_dictee",
    "t1_fin_dictee",
    "t2_affichage",
    "t3_premiere_decision",
    "t4_derniere_decision",
    "t5_cloture",
    "t6_export",
    "dictee_ms",
    "generation_ms",
    "revision_ms",
    "revision_nette_ms",
    "pauses_ms",
    "nb_pauses",
    # Les trois textes longs en dernier : places au milieu, ils rendraient le
    # fichier illisible des qu'on l'ouvre dans un tableur.
    "transcription",
    "cr_propose",
    "cr_valide",
)

COLONNES_PROPOSITIONS: Final[tuple[str, ...]] = (
    "proposition_id",
    "dossier_id",
    "praticien",
    "dossier_exclu",
    "prelevement_rang",
    "prelevement_libelle",
    "type",
    "sous_type",
    "valeur_proposee",
    "chemin",
    "positions",
    "confiance",
    "empan_debut",
    "empan_fin",
    "ancree",
    "longueur_mots",
    "affiche_a",
    "decide_a",
    "latence_ms",
    "hative",
    "justif_ouverte",
    "justif_duree_ms",
    "decision_changee_apres_justif",
    "decision",
    "valeur_retenue",
    "nature_correction",
    "cause_erreur",
)

COLONNES_QUESTIONNAIRES: Final[tuple[str, ...]] = (
    "reponse_id",
    "praticien",
    "questionnaire",
    "item",
    "valeur",
    "repondu_a",
    "dossier_id",
    "dossier_exclu",
)

COLONNES_REVISIONS: Final[tuple[str, ...]] = (
    "proposition_id",
    "dossier_id",
    "praticien",
    "dossier_exclu",
    "rang",
    "decision",
    "etat",
    "valeur_retenue",
    "nature_correction",
    "cause_erreur",
    "justif_ouverte",
    "decide_a",
)

COLONNES_PAUSES: Final[tuple[str, ...]] = (
    "pause_id",
    "dossier_id",
    "praticien",
    "dossier_exclu",
    "debut",
    "fin",
    "duree_ms",
    "cause",
)


# --- Le corpus lu en une fois ------------------------------------------------


@dataclass(frozen=True)
class Corpus:
    """Tout ce que l'export doit lire, charge en une seule passe.

    L'export doit voir la MEME base pour les quatre fichiers : un dossier
    exclu entre deux requetes produirait une archive qui se contredit d'un
    fichier a l'autre.
    """

    sessions: list[EtudeSession]
    dossiers: list[EtudeDossier]
    prelevements: list[EtudePrelevement]
    propositions: list[EtudeProposition]
    reponses: list[EtudeReponseQuestionnaire]
    pauses: list[EtudePause]


_Table = TypeVar(
    "_Table",
    EtudeSession,
    EtudeDossier,
    EtudePrelevement,
    EtudeProposition,
    EtudeReponseQuestionnaire,
    EtudePause,
)


async def _tout_lire(base: AsyncSession, modele: type[_Table]) -> list[_Table]:
    """Lit une table entiere. Requete explicite : pas de lazy-load en async."""
    resultat = await base.execute(select(modele))
    return list(resultat.scalars().all())


async def charger_corpus(base: AsyncSession) -> Corpus:
    """Charge les six tables dont les quatre fichiers sont tires."""
    return Corpus(
        sessions=await _tout_lire(base, EtudeSession),
        dossiers=await _tout_lire(base, EtudeDossier),
        prelevements=await _tout_lire(base, EtudePrelevement),
        propositions=await _tout_lire(base, EtudeProposition),
        reponses=await _tout_lire(base, EtudeReponseQuestionnaire),
        pauses=await _tout_lire(base, EtudePause),
    )


# --- Pseudonymisation --------------------------------------------------------


def _en_utc(moment: datetime) -> datetime:
    """Ramene un horodatage a UTC pour que deux sources se comparent.

    SQLite rend des datetimes naifs la ou PostgreSQL les rend aware. Comparer
    les deux leve TypeError, ce qui ferait echouer l'export entier au moment
    ou l'on trie les praticiens par ordre d'inclusion.
    """
    if moment.tzinfo is None:
        return moment.replace(tzinfo=timezone.utc)
    return moment.astimezone(timezone.utc)


def premiere_inclusion(corpus: Corpus) -> dict[str, datetime]:
    """Date du premier evenement de chaque praticien.

    Les questionnaires comptent au meme titre que les sessions : un praticien
    qui a repondu au questionnaire d'inclusion sans avoir encore ouvert de
    dossier existe deja dans l'etude, et doit avoir son identifiant.
    """
    evenements: list[tuple[str, datetime]] = [
        (session.praticien_id, session.debut) for session in corpus.sessions
    ]
    evenements += [
        (reponse.praticien_id, reponse.repondu_a) for reponse in corpus.reponses
    ]

    premiers: dict[str, datetime] = {}
    for praticien_id, moment in evenements:
        instant = _en_utc(moment)
        connu = premiers.get(praticien_id)
        if connu is None or instant < connu:
            premiers[praticien_id] = instant
    return premiers


def attribuer_pseudonymes(premiers: dict[str, datetime]) -> dict[str, str]:
    """Numerote les praticiens par ordre d'inclusion : praticien_01, 02...

    L'ordre d'inclusion est le seul critere qui ne bouge pas : un praticien
    garde son numero d'un export a l'autre, sinon on ne peut ni suivre un
    praticien dans le temps ni recouper deux exports. L'identifiant de compte
    departage les ex aequo, pour que deux exports de la meme base soient
    identiques au bit pres.
    """
    ordonnes = sorted(premiers.items(), key=lambda paire: (paire[1], paire[0]))
    return {
        praticien_id: f"praticien_{rang:02d}"
        for rang, (praticien_id, _) in enumerate(ordonnes, start=1)
    }


# --- Correspondances ---------------------------------------------------------


@dataclass(frozen=True)
class Reperes:
    """Les correspondances dont les quatre fichiers ont besoin.

    Calculees une fois : sans elles, chaque ligne fille refarait la remontee
    dossier -> session -> praticien.
    """

    pseudonyme_par_praticien: dict[str, str]
    pseudonyme_par_dossier: dict[str, str]
    exclu_par_dossier: dict[str, bool]
    prelevement_par_id: dict[str, EtudePrelevement]


def construire_reperes(corpus: Corpus) -> Reperes:
    """Assemble les correspondances a partir du corpus charge."""
    pseudonymes = attribuer_pseudonymes(premiere_inclusion(corpus))
    praticien_par_session = {
        session.id: session.praticien_id for session in corpus.sessions
    }
    return Reperes(
        pseudonyme_par_praticien=pseudonymes,
        pseudonyme_par_dossier={
            dossier.id: pseudonymes.get(
                praticien_par_session.get(dossier.session_id, ""), ""
            )
            for dossier in corpus.dossiers
        },
        exclu_par_dossier={dossier.id: dossier.exclu for dossier in corpus.dossiers},
        prelevement_par_id={
            prelevement.id: prelevement for prelevement in corpus.prelevements
        },
    )


# --- Conversion des valeurs en cellules --------------------------------------


def _texte(valeur: str | None) -> str:
    """Une chaine absente devient une cellule vide, pas la chaine 'None'."""
    return valeur if valeur is not None else ""


def _nombre(valeur: int | float | None) -> str:
    """Un nombre non mesure reste VIDE : ecrire 0 inventerait une observation."""
    return "" if valeur is None else str(valeur)


def _booleen(valeur: bool | None) -> str:
    """Un booleen non renseigne reste VIDE, et surtout pas 'false'.

    `omission_signalee` vaut None tant que le dossier n'est pas clos. Le sortir
    en 'false' compterait chaque dossier en cours comme un dossier sans
    omission, et le taux d'omission publie serait faux vers le bas.
    """
    if valeur is None:
        return ""
    return "true" if valeur else "false"


def _horodatage(moment: datetime | None) -> str:
    """ISO 8601, ou cellule vide quand l'etape n'a pas eu lieu."""
    return "" if moment is None else moment.isoformat()


# --- Regroupement des tables filles ------------------------------------------

_Fille = TypeVar("_Fille", EtudeProposition, EtudePause)


def _grouper_par_dossier(lignes: list[_Fille]) -> dict[str, list[_Fille]]:
    """Range les lignes filles sous leur dossier, en une passe."""
    groupes: dict[str, list[_Fille]] = {}
    for ligne in lignes:
        groupes.setdefault(ligne.dossier_id, []).append(ligne)
    return groupes


# --- dossiers.csv -------------------------------------------------------------


def ligne_dossier(
    dossier: EtudeDossier,
    reperes: Reperes,
    propositions: list[EtudeProposition],
    pauses: list[EtudePause],
) -> dict[str, str]:
    """Un dossier, ses temps, son praticien pseudonymise et ses trois textes."""
    temps = calculer_temps(
        dossier, sum(pause.duree_ms or 0 for pause in pauses), len(pauses)
    )
    return {
        "dossier_id": dossier.id,
        "session_id": dossier.session_id,
        "praticien": reperes.pseudonyme_par_dossier.get(dossier.id, ""),
        "index_session": _nombre(dossier.index_session),
        "cree_a": _horodatage(dossier.cree_a),
        "organe": _texte(dossier.organe),
        "exclu": _booleen(dossier.exclu),
        "motif_exclusion": _texte(dossier.motif_exclusion),
        "abandonne": _booleen(dossier.abandonne),
        "motif_abandon": _texte(dossier.motif_abandon),
        "omission_signalee": _booleen(dossier.omission_signalee),
        "omission_texte": _texte(dossier.omission_texte),
        "nb_prelevements_detecte": _nombre(dossier.nb_prelevements_detecte),
        "nb_prelevements_corrige": _nombre(dossier.nb_prelevements_corrige),
        "nb_propositions": _nombre(len(propositions)),
        "nb_decidees": _nombre(
            sum(1 for p in propositions if p.decision is not None)
        ),
        "caracteres_modifies": _nombre(dossier.caracteres_modifies),
        "t0_debut_dictee": _horodatage(dossier.t0_debut_dictee),
        "t1_fin_dictee": _horodatage(dossier.t1_fin_dictee),
        "t2_affichage": _horodatage(dossier.t2_affichage),
        "t3_premiere_decision": _horodatage(dossier.t3_premiere_decision),
        "t4_derniere_decision": _horodatage(dossier.t4_derniere_decision),
        "t5_cloture": _horodatage(dossier.t5_cloture),
        "t6_export": _horodatage(dossier.t6_export),
        "dictee_ms": _nombre(temps.dictee_ms),
        "generation_ms": _nombre(temps.generation_ms),
        "revision_ms": _nombre(temps.revision_ms),
        "revision_nette_ms": _nombre(temps.revision_nette_ms),
        "pauses_ms": _nombre(temps.pauses_ms),
        "nb_pauses": _nombre(temps.nb_pauses),
        "transcription": dossier.transcription,
        "cr_propose": dossier.cr_propose,
        "cr_valide": _texte(dossier.cr_valide),
    }


def lignes_dossiers(corpus: Corpus, reperes: Reperes) -> list[dict[str, str]]:
    """Un dossier par ligne, EXCLUS COMPRIS.

    Les dossiers ecartes partent avec leur colonne `exclu` et leur motif : sans
    eux, l'effectif ecarte devient invisible et le diagramme de flux de la
    publication ne peut plus etre rempli.
    """
    par_dossier_propositions = _grouper_par_dossier(corpus.propositions)
    par_dossier_pauses = _grouper_par_dossier(corpus.pauses)
    ordonnes = sorted(
        corpus.dossiers,
        key=lambda d: (
            reperes.pseudonyme_par_dossier.get(d.id, ""),
            d.session_id,
            d.index_session,
            d.id,
        ),
    )
    return [
        ligne_dossier(
            dossier,
            reperes,
            par_dossier_propositions.get(dossier.id, []),
            par_dossier_pauses.get(dossier.id, []),
        )
        for dossier in ordonnes
    ]


# --- propositions.csv ---------------------------------------------------------


def ligne_proposition(
    proposition: EtudeProposition, reperes: Reperes
) -> dict[str, str]:
    """Une proposition, sa decision, et le prelevement qu'elle vise."""
    prelevement = (
        reperes.prelevement_par_id.get(proposition.prelevement_id)
        if proposition.prelevement_id is not None
        else None
    )
    return {
        "proposition_id": proposition.id,
        "dossier_id": proposition.dossier_id,
        "praticien": reperes.pseudonyme_par_dossier.get(proposition.dossier_id, ""),
        # Repete ici pour que le filtre des exclus se fasse sans jointure :
        # une jointure oubliee est une jointure qui n'aura pas lieu.
        "dossier_exclu": _booleen(
            reperes.exclu_par_dossier.get(proposition.dossier_id)
        ),
        "prelevement_rang": _nombre(prelevement.rang if prelevement else None),
        "prelevement_libelle": _texte(prelevement.libelle if prelevement else None),
        "type": proposition.type,
        "sous_type": _texte(proposition.sous_type),
        "valeur_proposee": proposition.valeur_proposee,
        "chemin": _texte(proposition.chemin),
        "positions": _texte(proposition.positions),
        "confiance": _nombre(proposition.confiance),
        "empan_debut": _nombre(proposition.empan_debut),
        "empan_fin": _nombre(proposition.empan_fin),
        "ancree": _booleen(proposition.empan_debut is not None),
        "longueur_mots": _nombre(proposition.longueur_mots),
        "affiche_a": _horodatage(proposition.affiche_a),
        "decide_a": _horodatage(proposition.decide_a),
        "latence_ms": _nombre(proposition.latence_ms),
        "hative": _booleen(proposition.hative),
        "justif_ouverte": _booleen(proposition.justif_ouverte),
        "justif_duree_ms": _nombre(proposition.justif_duree_ms),
        "decision_changee_apres_justif": _booleen(
            proposition.decision_changee_apres_justif
        ),
        "decision": _texte(proposition.decision),
        "valeur_retenue": _texte(proposition.valeur_retenue),
        "nature_correction": _texte(proposition.nature_correction),
        "cause_erreur": _texte(proposition.cause_erreur),
    }


def lignes_propositions(corpus: Corpus, reperes: Reperes) -> list[dict[str, str]]:
    """Une proposition par ligne, avec sa decision."""
    ordonnees = sorted(corpus.propositions, key=lambda p: (p.dossier_id, p.id))
    return [ligne_proposition(proposition, reperes) for proposition in ordonnees]


# --- questionnaires.csv -------------------------------------------------------


def ligne_questionnaire(
    reponse: EtudeReponseQuestionnaire, reperes: Reperes
) -> dict[str, str]:
    """Une reponse a un item, rattachee a son praticien pseudonymise."""
    return {
        "reponse_id": reponse.id,
        "praticien": reperes.pseudonyme_par_praticien.get(reponse.praticien_id, ""),
        "questionnaire": reponse.questionnaire,
        "item": reponse.item,
        "valeur": reponse.valeur,
        "repondu_a": _horodatage(reponse.repondu_a),
        "dossier_id": _texte(reponse.dossier_id),
        # Vide, et non 'false', quand la reponse ne porte sur aucun dossier :
        # les questionnaires d'inclusion et de fin d'etude n'en visent aucun.
        "dossier_exclu": _booleen(
            reperes.exclu_par_dossier.get(reponse.dossier_id)
            if reponse.dossier_id is not None
            else None
        ),
    }


def lignes_revisions(corpus: Corpus, reperes: Reperes) -> list[dict[str, str]]:
    """TOUTE decision jamais prise, y compris celles remplacees depuis.

    `propositions.csv` porte l'etat FINAL : c'est lui qui sert les taux, et il
    ne doit pas bouger. Cette table-ci porte le chemin parcouru pour y arriver.
    Une proposition decidee une seule fois y occupe une ligne ; une proposition
    sur laquelle le praticien est revenu en occupe autant que d'avis.

    Sans elle, un clic errant rattrape dans la seconde et un revirement apres
    lecture de la justification sont indistinguables au depouillement.
    """
    lignes: list[dict[str, str]] = []
    for proposition in sorted(corpus.propositions, key=lambda p: (p.dossier_id, p.id)):
        for revision in proposition.revisions:
            lignes.append(
                {
                    "proposition_id": proposition.id,
                    "dossier_id": proposition.dossier_id,
                    "praticien": reperes.pseudonyme_par_dossier.get(
                        proposition.dossier_id, ""
                    ),
                    "dossier_exclu": _booleen(
                        reperes.exclu_par_dossier.get(proposition.dossier_id)
                    ),
                    "rang": _nombre(revision.rang),
                    "decision": revision.decision,
                    "etat": revision.etat,
                    "valeur_retenue": _texte(revision.valeur_retenue),
                    "nature_correction": _texte(revision.nature_correction),
                    "cause_erreur": _texte(revision.cause_erreur),
                    "justif_ouverte": _booleen(revision.justif_ouverte),
                    "decide_a": _horodatage(revision.decide_a),
                }
            )
    return lignes


def lignes_questionnaires(corpus: Corpus, reperes: Reperes) -> list[dict[str, str]]:
    """Une reponse par ligne."""
    ordonnees = sorted(
        corpus.reponses,
        key=lambda r: (
            reperes.pseudonyme_par_praticien.get(r.praticien_id, ""),
            r.questionnaire,
            r.item,
            r.id,
        ),
    )
    return [ligne_questionnaire(reponse, reperes) for reponse in ordonnees]


# --- pauses.csv ---------------------------------------------------------------


def ligne_pause(pause: EtudePause, reperes: Reperes) -> dict[str, str]:
    """Une interruption du chronometre."""
    return {
        "pause_id": pause.id,
        "dossier_id": pause.dossier_id,
        "praticien": reperes.pseudonyme_par_dossier.get(pause.dossier_id, ""),
        "dossier_exclu": _booleen(reperes.exclu_par_dossier.get(pause.dossier_id)),
        "debut": _horodatage(pause.debut),
        "fin": _horodatage(pause.fin),
        "duree_ms": _nombre(pause.duree_ms),
        "cause": pause.cause,
    }


def lignes_pauses(corpus: Corpus, reperes: Reperes) -> list[dict[str, str]]:
    """Une pause par ligne."""
    ordonnees = sorted(corpus.pauses, key=lambda p: (p.dossier_id, p.id))
    return [ligne_pause(pause, reperes) for pause in ordonnees]


# --- Serialisation ------------------------------------------------------------


def ecrire_csv(colonnes: tuple[str, ...], lignes: list[dict[str, str]]) -> bytes:
    """Serialise un tableau en CSV UTF-8 avec BOM.

    Le module csv de la bibliotheque standard, jamais une concatenation : la
    transcription et les deux comptes rendus contiennent des retours a la ligne
    et des guillemets, qu'il faut echapper pour que le fichier reste lisible.

    Le BOM n'est pas cosmetique : sans lui, Excel lit l'UTF-8 comme du Latin-1
    et rend les accents illisibles — or Excel sera le premier lecteur de ces
    fichiers.
    """
    tampon = io.StringIO(newline="")
    graveur = csv.DictWriter(
        tampon, fieldnames=list(colonnes), restval="", extrasaction="raise"
    )
    graveur.writeheader()
    graveur.writerows(lignes)
    return tampon.getvalue().encode("utf-8-sig")


@dataclass(frozen=True)
class Fichier:
    """Un fichier de l'archive et son effectif.

    Le nombre de lignes voyage avec le contenu : le lisez-moi doit annoncer un
    effectif verifiable, pas un effectif recompte a la main.
    """

    nom: str
    contenu: bytes
    nb_lignes: int


def construire_fichiers(corpus: Corpus) -> list[Fichier]:
    """Produit les cinq CSV a partir du corpus charge."""
    reperes = construire_reperes(corpus)
    tables: tuple[tuple[str, tuple[str, ...], list[dict[str, str]]], ...] = (
        (NOM_DOSSIERS, COLONNES_DOSSIERS, lignes_dossiers(corpus, reperes)),
        (
            NOM_PROPOSITIONS,
            COLONNES_PROPOSITIONS,
            lignes_propositions(corpus, reperes),
        ),
        (
            NOM_QUESTIONNAIRES,
            COLONNES_QUESTIONNAIRES,
            lignes_questionnaires(corpus, reperes),
        ),
        (NOM_PAUSES, COLONNES_PAUSES, lignes_pauses(corpus, reperes)),
        (NOM_REVISIONS, COLONNES_REVISIONS, lignes_revisions(corpus, reperes)),
    )
    return [
        Fichier(nom, ecrire_csv(colonnes, lignes), len(lignes))
        for nom, colonnes, lignes in tables
    ]


# --- Dictionnaire de donnees --------------------------------------------------
#
# Un export sans dictionnaire est inexploitable six mois plus tard : le
# statisticien qui ouvre l'archive n'aura pas le code sous les yeux, et
# `hative` ou `nature_correction` ne se devinent pas.

_DEFINITIONS: Final[tuple[tuple[str, tuple[tuple[str, str], ...]], ...]] = (
    (
        NOM_DOSSIERS,
        (
            ("praticien", "Pseudonyme stable, attribue par ordre d'inclusion."),
            ("index_session", "Rang du cas dans la session : 0 pour le premier."),
            (
                "exclu",
                "Dossier ecarte de l'analyse (essai, saisie aberrante). "
                "Il doit etre RETIRE de tout calcul de taux ; il est exporte "
                "pour que l'effectif ecarte soit rendu public.",
            ),
            ("motif_exclusion", "Raison de l'exclusion, telle que saisie."),
            ("abandonne", "Le praticien a quitte le cas sans le clore."),
            (
                "omission_signalee",
                "Le praticien a signale un element dicte et absent du CR. "
                "Vide tant que le dossier n'est pas clos : ce vide n'est PAS "
                "une absence d'omission.",
            ),
            ("t0_debut_dictee", "Debut de la dictee."),
            ("t1_fin_dictee", "Fin de la dictee."),
            ("t2_affichage", "Affichage du compte rendu propose."),
            ("t3_premiere_decision", "Premiere decision prise sur une proposition."),
            ("t4_derniere_decision", "Derniere decision prise."),
            ("t5_cloture", "Cloture du dossier."),
            ("t6_export", "Export du compte rendu par le praticien."),
            ("dictee_ms", "t1 - t0."),
            ("generation_ms", "t2 - t1."),
            (
                "revision_ms",
                "t2 - t5. C'est le referentiel de temps de l'etude : la "
                "redaction AVEC l'outil, telle que le praticien la vit.",
            ),
            ("revision_nette_ms", "revision_ms moins pauses_ms, plancher a zero."),
            (
                "pauses_ms",
                "Somme des pauses CLOSES du dossier. Une pause encore ouverte "
                "compte dans nb_pauses mais pas dans cette somme.",
            ),
            (
                "caracteres_modifies",
                "Ecart entre cr_propose et cr_valide : la charge d'edition.",
            ),
            ("cr_propose", "Ce que le systeme a propose, fige a l'affichage."),
            ("cr_valide", "Ce que le praticien a valide."),
        ),
    ),
    (
        NOM_PROPOSITIONS,
        (
            (
                "dossier_exclu",
                "Recopie de dossiers.csv/exclu, pour filtrer sans jointure.",
            ),
            (
                "type",
                "restitution (ce que le systeme a compris de la dictee), "
                "code (ADICAP) ou completude (element possiblement manquant).",
            ),
            (
                "decision",
                "Trois grilles distinctes, a ne jamais confondre. "
                "restitution : conforme / corrige / non_dicte / hors_sujet. "
                "code : juste / corrige / je_ne_sais_pas. "
                "completude : pertinent_ajoute / pertinent_non_retenu / "
                "non_pertinent.",
            ),
            (
                "ancree",
                "false = aucun passage de la dictee ne soutient la "
                "proposition. C'est la mesure d'hallucination cote systeme ; "
                "non_dicte est la meme mesure cote praticien.",
            ),
            ("empan_debut", "Offset de caractere dans dossiers.csv/transcription."),
            ("empan_fin", "Offset de fin, exclu."),
            ("latence_ms", "Delai entre affiche_a et decide_a."),
            (
                "hative",
                "Decision plus rapide que 1200 ms sur une proposition de plus "
                "de 15 mots, donc trop rapide pour avoir ete lue. A analyser "
                "a part, jamais a supprimer.",
            ),
            (
                "nature_correction",
                "style (le fond etait juste, le praticien reformule), "
                "precision (juste mais incomplet), erreur_fond (le fond etait "
                "FAUX). Seul erreur_fond impute une erreur au systeme.",
            ),
            (
                "cause_erreur",
                "transcription ou interpretation. Ne se pose que sur une "
                "erreur de fond.",
            ),
            (
                "decision_changee_apres_justif",
                "Le praticien a change d'avis apres avoir ouvert la "
                "justification : la mesure d'explicabilite.",
            ),
        ),
    ),
    (
        NOM_QUESTIONNAIRES,
        (
            (
                "questionnaire",
                "inclusion, par_cas, periodique (le F-SUS, tous les 5 "
                "comptes rendus clos) ou fin_etude.",
            ),
            ("item", "Identifiant stable de l'item, ex fsus_03, par_cas_04."),
            ("valeur", "Reponse brute, telle que saisie."),
            (
                "dossier_id",
                "Vide pour les questionnaires d'inclusion et de fin d'etude, "
                "qui ne portent sur aucun cas.",
            ),
        ),
    ),
    (
        NOM_PAUSES,
        (
            (
                "cause",
                "onglet_masque (onglet passe en arriere-plan) ou inactivite "
                "(plus de 90 s sans evenement).",
            ),
            ("fin", "Vide si la pause n'a jamais ete refermee."),
            ("duree_ms", "Vide si la pause n'a jamais ete refermee."),
        ),
    ),
    (
        NOM_REVISIONS,
        (
            (
                "rang",
                "1 pour la premiere decision prise sur le bloc, 2 pour la "
                "suivante. Le rang le plus eleve d'une proposition correspond "
                "a la decision publiee dans propositions.csv.",
            ),
            (
                "decision",
                "L'avis a CE moment-la, y compris s'il a ete remplace depuis. "
                "NE PAS agreger cette table pour calculer un taux : elle "
                "compterait plusieurs fois les blocs sur lesquels le praticien "
                "est revenu. Les taux se calculent sur propositions.csv, qui "
                "porte un etat final par bloc.",
            ),
            (
                "justif_ouverte",
                "La justification etait-elle ouverte au moment de CETTE "
                "decision. Comparer ce champ entre deux rangs successifs "
                "permet de demontrer qu'une explication a change un avis, au "
                "lieu de s'en remettre au seul booleen agrege.",
            ),
            (
                "decide_a",
                "Horodatage de cette decision-la. L'ecart entre deux rangs "
                "separe un clic errant rattrape dans la seconde d'un "
                "revirement apres relecture.",
            ),
        ),
    ),
)


_TITRE: Final = "EXPORT DE L'ETUDE MARC\n======================"

#: Largeur de mise en page du lisez-moi. Les descriptions sont repliees a la
#: main : un dictionnaire de donnees qui deborde de l'ecran ne se lit pas, et
#: c'est celui-la qu'on ouvrira six mois plus tard sans avoir le code sous
#: les yeux.
_LARGEUR: Final[int] = 78

_CONVENTIONS: Final = """CONVENTIONS COMMUNES A TOUS LES FICHIERS
----------------------------------------
- CSV encode en UTF-8 avec BOM, separateur virgule, guillemets doubles
  (RFC 4180). Les colonnes de texte contiennent des retours a la ligne :
  ouvrez-les avec un vrai lecteur de CSV (read.csv, pandas, un tableur), et
  jamais avec un decoupage sur la virgule.
- UNE CELLULE VIDE SIGNIFIE "NON MESURE". Jamais zero, jamais faux. Remplacer
  ces vides par 0 avant de calculer une moyenne, ou par false avant de
  calculer un taux, produirait un chiffre faux et invisible.
- Les booleens s'ecrivent true / false.
- Les horodatages sont au format ISO 8601. Toutes les durees sont en
  millisecondes et se terminent par _ms.
- L'ordre des colonnes est fige : deux exports successifs se comparent
  colonne par colonne.

PSEUDONYMISATION
----------------
Les praticiens apparaissent sous un identifiant technique (praticien_01,
praticien_02...) attribue par ordre d'inclusion dans l'etude. Cet identifiant
est STABLE d'un export a l'autre : un meme praticien porte toujours le meme
numero, ce qui permet de le suivre dans le temps et de recouper deux exports.
Aucun nom, aucune adresse electronique, aucun identifiant de compte ne figure
dans ces fichiers.

DOSSIERS EXCLUS
---------------
Les dossiers ecartes de l'etude (essais, saisies aberrantes) SONT exportes,
avec leur colonne `exclu` et leur motif : sans eux, l'effectif ecarte devient
invisible, alors que toute publication demande de le declarer. C'est a
l'analyse de filtrer, pas a l'export de cacher. Le drapeau est recopie sous le
nom `dossier_exclu` dans propositions.csv, questionnaires.csv et pauses.csv,
pour que le filtre se fasse sans jointure.

LIAISONS ENTRE FICHIERS
-----------------------
dossiers.csv/dossier_id est la cle. propositions.csv, questionnaires.csv et
pauses.csv la referencent. dossiers.csv/session_id regroupe les cas d'une meme
seance, dans l'ordre donne par index_session."""


def _section_contenu(fichiers: list[Fichier], moment: datetime) -> str:
    """Date de l'export et effectif de chaque fichier."""
    lignes = [
        f"Export genere le {moment.isoformat()}.",
        "",
        "CONTENU",
        "-------",
    ]
    largeur = max(len(fichier.nom) for fichier in fichiers)
    for fichier in fichiers:
        lignes.append(
            f"  {fichier.nom:<{largeur}}  {fichier.nb_lignes} ligne(s) hors en-tete"
        )
    lignes.append(f"  {NOM_LISEZ_MOI:<{largeur}}  ce fichier")
    return "\n".join(lignes)


def _definition(colonne: str, description: str) -> str:
    """Met en page une entree du dictionnaire, repliee a la largeur de l'ecran."""
    corps = textwrap.fill(
        description,
        width=_LARGEUR,
        initial_indent="      ",
        subsequent_indent="      ",
    )
    return f"  {colonne}\n{corps}"


def _section_dictionnaire() -> str:
    """Signification des colonnes qui ne se devinent pas."""
    lignes = ["DICTIONNAIRE DES COLONNES", "-------------------------"]
    for nom_fichier, definitions in _DEFINITIONS:
        lignes.append("")
        lignes.append(nom_fichier)
        lignes += [
            _definition(colonne, description) for colonne, description in definitions
        ]
    return "\n".join(lignes)


def rediger_lisez_moi(fichiers: list[Fichier], moment: datetime) -> str:
    """Assemble le dictionnaire de donnees qui accompagne l'archive."""
    return "\n\n".join(
        (
            _TITRE,
            _section_contenu(fichiers, moment),
            _CONVENTIONS,
            _section_dictionnaire(),
        )
    ) + "\n"


# --- Archive ------------------------------------------------------------------


def construire_archive(corpus: Corpus, moment: datetime) -> bytes:
    """Assemble les quatre CSV et leur lisez-moi en une archive ZIP."""
    fichiers = construire_fichiers(corpus)
    tampon = io.BytesIO()
    with zipfile.ZipFile(tampon, "w", zipfile.ZIP_DEFLATED) as archive:
        for fichier in fichiers:
            archive.writestr(fichier.nom, fichier.contenu)
        archive.writestr(
            NOM_LISEZ_MOI, rediger_lisez_moi(fichiers, moment).encode("utf-8")
        )
    return tampon.getvalue()


def construire_classeur(corpus: Corpus, moment: datetime) -> bytes:
    """Assemble les memes tables en un classeur Excel, un onglet par table.

    POURQUOI EN PLUS DU CSV, et pas a la place.

    Le CSV est le format d'ANALYSE : il se lit dans R, en Python, sans rien
    installer, et il ne deforme rien. Le classeur est le format de TRAVAIL :
    c'est celui qu'on ouvre pour regarder, trier, montrer a quelqu'un. Les deux
    repondent a des besoins differents, et n'en servir qu'un ferait perdre l'un
    des deux publics.

    Le classeur porte les MEMES colonnes et les MEMES lignes que les CSV : deux
    exports du meme corpus doivent se recouper a la ligne pres, sinon on ne sait
    plus lequel fait foi.

    Le lisez-moi devient un onglet : dans un classeur, un fichier texte a cote
    ne serait jamais ouvert.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font

    reperes = construire_reperes(corpus)
    tables = (
        (NOM_DOSSIERS, COLONNES_DOSSIERS, lignes_dossiers(corpus, reperes)),
        (NOM_PROPOSITIONS, COLONNES_PROPOSITIONS, lignes_propositions(corpus, reperes)),
        (NOM_QUESTIONNAIRES, COLONNES_QUESTIONNAIRES, lignes_questionnaires(corpus, reperes)),
        (NOM_PAUSES, COLONNES_PAUSES, lignes_pauses(corpus, reperes)),
        (NOM_REVISIONS, COLONNES_REVISIONS, lignes_revisions(corpus, reperes)),
    )

    classeur = Workbook()
    classeur.remove(classeur.active)

    for nom, colonnes, lignes in tables:
        onglet = classeur.create_sheet(Path(nom).stem[:31])
        onglet.append(list(colonnes))
        for cellule in onglet[1]:
            cellule.font = Font(bold=True)
        for ligne in lignes:
            onglet.append([ligne.get(colonne, "") for colonne in colonnes])
        # Les en-tetes restent visibles au defilement : sans cela, une table de
        # milliers de lignes devient illisible des le premier ecran.
        onglet.freeze_panes = "A2"
        for index, colonne in enumerate(colonnes, start=1):
            largeur = max(len(colonne), 12)
            onglet.column_dimensions[onglet.cell(1, index).column_letter].width = min(
                largeur + 2, 60
            )

    lecture = classeur.create_sheet("lisez-moi", 0)
    for numero, ligne in enumerate(
        rediger_lisez_moi(construire_fichiers(corpus), moment).splitlines(), start=1
    ):
        lecture.cell(numero, 1, ligne).alignment = Alignment(wrap_text=False)
    lecture.column_dimensions["A"].width = 110

    tampon = io.BytesIO()
    classeur.save(tampon)
    return tampon.getvalue()


def nom_classeur(moment: datetime) -> str:
    """Nomme le classeur par sa date, a la seconde, comme l'archive."""
    return f"export_etude_marc_{moment.strftime('%Y%m%d_%H%M%S')}.xlsx"


def nom_archive(moment: datetime) -> str:
    """Nomme l'archive par sa date, a la seconde.

    Deux exports du meme jour tomberaient sinon sur le meme nom de fichier, et
    le second ecraserait le premier dans le dossier de telechargement.
    """
    return f"export_etude_marc_{moment.strftime('%Y%m%d_%H%M%S')}.zip"
